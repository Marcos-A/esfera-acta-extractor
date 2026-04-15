#!/usr/bin/env python3
"""
Compare two generated workbooks for layout and cell-value differences.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def compare_workbooks(left_path: Path, right_path: Path, diff_limit: int) -> dict[str, object]:
    """Compare the first worksheet of two exported workbooks for regression checks."""
    left_wb = load_workbook(left_path)
    right_wb = load_workbook(right_path)
    left_ws = left_wb.active
    right_ws = right_wb.active

    header_left = [left_ws.cell(row=1, column=index).value for index in range(1, left_ws.max_column + 1)]
    header_right = [right_ws.cell(row=1, column=index).value for index in range(1, right_ws.max_column + 1)]

    diffs: list[dict[str, object]] = []
    max_row = max(left_ws.max_row, right_ws.max_row)
    max_col = max(left_ws.max_column, right_ws.max_column)
    # Stop after diff_limit mismatches so the output stays readable during quick checks.
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            left_value = left_ws.cell(row=row, column=col).value
            right_value = right_ws.cell(row=row, column=col).value
            if left_value != right_value:
                diffs.append(
                    {
                        "row": row,
                        "column": col,
                        "left": left_value,
                        "right": right_value,
                    }
                )
                if len(diffs) >= diff_limit:
                    break
        if len(diffs) >= diff_limit:
            break

    return {
        "left": str(left_path),
        "right": str(right_path),
        "same_sheet_title": left_ws.title == right_ws.title,
        "same_dimensions": [left_ws.max_row, left_ws.max_column] == [right_ws.max_row, right_ws.max_column],
        "same_headers": header_left == header_right,
        "left_dimensions": [left_ws.max_row, left_ws.max_column],
        "right_dimensions": [right_ws.max_row, right_ws.max_column],
        "left_conditional_formatting_ranges": len(left_ws.conditional_formatting),
        "right_conditional_formatting_ranges": len(right_ws.conditional_formatting),
        "diffs": diffs,
    }


def main() -> None:
    """CLI entrypoint for quick workbook diffing during performance or refactor work."""
    parser = argparse.ArgumentParser(description="Compare two XLSX workbooks.")
    parser.add_argument("left", type=Path)
    parser.add_argument("right", type=Path)
    parser.add_argument("--diff-limit", type=int, default=20)
    args = parser.parse_args()

    result = compare_workbooks(args.left, args.right, diff_limit=args.diff_limit)
    print(json.dumps(result, ensure_ascii=True, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
