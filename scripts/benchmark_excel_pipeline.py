#!/usr/bin/env python3
"""
Benchmark the Excel export hot path using a synthetic dataset derived from a sample workbook.
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import statistics
import sys
import tempfile
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from src.excel_processor import export_excel_with_spacing


DEFAULT_SAMPLE_WORKBOOK = Path(
    "failed_uploads/e2312b3e90114580b07d57c8cf465503/output/CFPM_IC10_1_CFPM_IC10101.xlsx"
)


def _normalize_header(header: object) -> str:
    return str(header or "").replace("\n", "").strip()


def _parse_sample_workbook(sample_workbook: Path) -> tuple[list[str], list[str], dict[str, str], list[str]]:
    wb = load_workbook(sample_workbook)
    ws = wb.active
    headers = [_normalize_header(ws.cell(row=1, column=index).value) for index in range(1, ws.max_column + 1)]

    ra_headers: list[str] = []
    mp_codes: list[str] = []
    mp_codes_with_em: list[str] = []
    em_headers_by_mp: dict[str, str] = {}

    for header in headers[2:]:
        if not header:
            continue
        if header.endswith("RA"):
            ra_headers.append(header)
            continue
        if header.endswith(" CENTRE"):
            continue
        if header.endswith(" EMPRESA"):
            mp_code = header.split()[0]
            mp_codes_with_em.append(mp_code)
            continue
        mp_codes.append(header)

    for mp_code in mp_codes_with_em:
        matching_ra = next((header for header in ra_headers if header.startswith(f"{mp_code}_")), None)
        em_headers_by_mp[mp_code] = (
            matching_ra[:-2] + "EM" if matching_ra and matching_ra.endswith("RA") else f"{mp_code}_AUTO_01EM"
        )

    return ra_headers, mp_codes, em_headers_by_mp, headers


def _build_synthetic_dataframe(
    sample_workbook: Path,
    student_count: int,
) -> tuple[pd.DataFrame, list[str], list[str]]:
    ra_headers, mp_codes, em_headers_by_mp, _headers = _parse_sample_workbook(sample_workbook)

    rows: list[dict[str, object]] = []
    numeric_cycle = [7, 8, 4, 9, 6]
    status_cycle = ["PDT", "EP", "", "PQ"]

    for index in range(student_count):
        row: dict[str, object] = {"estudiant": f"Student {index + 1:03d}"}
        for ra_index, header in enumerate(ra_headers):
            cycle_value = numeric_cycle[(index + ra_index) % len(numeric_cycle)]
            row[header] = cycle_value if (index + ra_index) % 6 else status_cycle[(index + ra_index) % len(status_cycle)]
        for mp_code in mp_codes:
            row[mp_code] = numeric_cycle[(index + len(mp_code)) % len(numeric_cycle)]
        for mp_code, em_header in em_headers_by_mp.items():
            row[em_header] = numeric_cycle[(index + len(mp_code) + 1) % len(numeric_cycle)]
        rows.append(row)

    df = pd.DataFrame(rows)
    return df, sorted(em_headers_by_mp), mp_codes


def benchmark(sample_workbook: Path, student_count: int, runs: int) -> dict[str, object]:
    df, _em_headers, mp_codes = _build_synthetic_dataframe(sample_workbook, student_count=student_count)

    mp_codes_with_em = sorted(
        {
            column.split("_", 1)[0]
            for column in df.columns
            if column.endswith("EM")
        }
    )

    output_dir = Path(tempfile.mkdtemp(prefix="esfera-benchmark-"))
    os.environ["PERF_TIMING_ENABLED"] = "1"
    durations: list[float] = []

    try:
        for run_index in range(1, runs + 1):
            output_path = output_dir / f"benchmark-run-{run_index}.xlsx"
            started_at = time.perf_counter()
            export_excel_with_spacing(df.copy(), str(output_path), mp_codes_with_em, mp_codes)
            durations.append(time.perf_counter() - started_at)
    finally:
        shutil.rmtree(output_dir, ignore_errors=True)

    return {
        "sample_workbook": str(sample_workbook),
        "student_count": student_count,
        "runs": runs,
        "seconds": [round(duration, 6) for duration in durations],
        "mean_seconds": round(statistics.mean(durations), 6),
        "min_seconds": round(min(durations), 6),
        "max_seconds": round(max(durations), 6),
        "dataframe_shape": list(df.shape),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Benchmark the Excel export path.")
    parser.add_argument("--sample-workbook", type=Path, default=DEFAULT_SAMPLE_WORKBOOK)
    parser.add_argument("--students", type=int, default=60)
    parser.add_argument("--runs", type=int, default=3)
    args = parser.parse_args()

    if not args.sample_workbook.exists():
        raise SystemExit(f"Sample workbook not found: {args.sample_workbook}")

    result = benchmark(args.sample_workbook, student_count=args.students, runs=args.runs)
    print(json.dumps(result, ensure_ascii=True, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
