"""
Excel processing module for generating and formatting grade reports.
"""

import json
from functools import lru_cache
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from typing import Optional

from .perf_timing import TimingRecorder


@lru_cache(maxsize=1)
def _load_mp_name_lookup() -> dict[str, str]:
    """Load the bundled MP-code to friendly-name mapping."""
    mapping_path = Path(__file__).resolve().parent / 'data' / 'mp_code_names.json'
    payload = json.loads(mapping_path.read_text(encoding='utf-8'))
    lookup: dict[str, str] = {}
    for modules in payload.values():
        for module in modules:
            code = module.get('codi')
            name = module.get('nom')
            if code and name and code not in lookup:
                lookup[code] = name
    return lookup


def _format_mp_display_name(mp_code: str) -> str:
    """Format a summary label as 'Code. Friendly name' when known."""
    friendly_name = _load_mp_name_lookup().get(mp_code)
    if friendly_name:
        return f'{mp_code}. {friendly_name}'
    return mp_code


def apply_row_formatting(
    workbook_path: str,
    mp_codes_with_em: list[str],
    mp_codes: list[str],
) -> None:
    """Apply the visual layout that staff expects in the exported workbook."""
    wb = load_workbook(workbook_path)
    ws = wb.active
    _apply_row_formatting_to_sheet(ws, mp_codes_with_em, mp_codes)
    wb.save(workbook_path)

def apply_conditional_formatting(
    workbook_path: str,
    mp_groups: dict[str, list[str]],
    mp_codes_with_em: list[str],
    mp_codes: list[str],
) -> None:
    """Highlight values that need review, such as missing grades or failing marks."""
    wb = load_workbook(workbook_path)
    ws = wb.active
    _apply_conditional_formatting_to_sheet(ws, mp_groups, mp_codes_with_em, mp_codes)
    wb.save(workbook_path)


def export_excel_with_spacing(
    df: pd.DataFrame,
    output_path: str,
    mp_codes_with_em: list[str],
    mp_codes: list[str],
    include_summary_sheet: bool = False,
) -> None:
    """
    Export DataFrame to Excel with specific column spacing after each MP's RAs.
    Adds a "#" column with sequential numbers at the beginning.
    """
    timings = TimingRecorder("export_excel_with_spacing")
    # MP grade columns are renamed to the short MP code, while detailed RA/EM columns
    # keep their full identifier so the final workbook remains understandable.
    with timings.measure("prepare_export_dataframe"):
        non_mp_columns = [col for col in df.columns
                             if col.endswith(('EM', 'RA')) or col == 'estudiant']
        df = df.rename(columns=lambda col: col.split('_')[0] if col not in non_mp_columns else col)
        mp_grade_columns = [col for col in df.columns if not col.endswith('EM') and
                             not col.endswith('RA') and col != 'estudiant']
        df_without_mp_grades = df[non_mp_columns].copy()

        ra_codes = [col for col in df_without_mp_grades.columns if col.endswith('RA')]

        mp_groups = {mp: [] for mp in mp_codes}
        sorted_mp_codes = sorted(mp_codes, key=len, reverse=True)
        for ra_code in ra_codes:
            for mp_code in sorted_mp_codes:
                if ra_code.startswith(mp_code + '_'):
                    mp_groups[mp_code].append(ra_code)
                    break

        new_columns = ['estudiant']
        for mp_code in mp_codes:
            # Keep RA columns grouped under their parent MP so the final workbook reads
            # left-to-right the way staff expect when reviewing one module at a time.
            for ra in mp_groups[mp_code]:
                new_columns.append(ra)
            if mp_code in mp_codes_with_em:
                new_columns.extend([
                    f'{mp_code} CENTRE',
                    f'{mp_code} EMPRESA',
                    f'{mp_code}'
                ])
            else:
                new_columns.append(f'{mp_code}')

        em_to_mp = {}
        for mp_code in mp_codes_with_em:
            em_codes = [col for col in df_without_mp_grades.columns if col.startswith(f'{mp_code}_') and col.endswith('EM')]
            for em_code in em_codes:
                em_to_mp[em_code] = mp_code

        export_df = df_without_mp_grades.reindex(columns=new_columns)
        for em_code, mp_code in em_to_mp.items():
            empresa_header = f'{mp_code} EMPRESA'
            if empresa_header in export_df.columns and em_code in df_without_mp_grades.columns:
                export_df[empresa_header] = df_without_mp_grades[em_code]

        for mp_code in mp_grade_columns:
            if mp_code in export_df.columns:
                export_df[mp_code] = df[mp_code]

        export_df = _blank_literal_na(export_df)
        object_columns = export_df.select_dtypes(include=['object', 'string']).columns
        if len(object_columns) > 0:
            export_df[object_columns] = export_df[object_columns].fillna("")

        # Add sequential index column
        export_df.insert(0, '#', range(1, len(export_df) + 1))

    output_path = output_path.replace('.csv', '.xlsx')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        with timings.measure("initial_to_excel"):
            export_df.to_excel(writer, index=False)
        ws = writer.sheets[next(iter(writer.sheets))]
        ws.title = 'Acta'
        with timings.measure("apply_row_formatting"):
            _apply_row_formatting_to_sheet(ws, mp_codes_with_em, mp_codes)
        with timings.measure("apply_conditional_formatting"):
            _apply_conditional_formatting_to_sheet(ws, mp_groups, mp_codes_with_em, mp_codes)
        if include_summary_sheet:
            with timings.measure("append_summary_sheet"):
                _append_pending_ra_summary_sheet(writer, df)

    timings.log(
        output_path=output_path,
        student_rows=len(export_df),
        output_columns=len(export_df.columns),
        mp_count=len(mp_codes),
        mp_with_em_count=len(mp_codes_with_em),
    )
    print(f"\t- Exported {len(export_df)} entries to {output_path}")


def _blank_literal_na(df: pd.DataFrame) -> pd.DataFrame:
    """Clear text cells containing the literal marker 'NA' before exporting to Excel."""
    cleaned_df = df.copy()
    object_columns = cleaned_df.select_dtypes(include=['object', 'string']).columns
    if len(object_columns) > 0:
        cleaned_df[object_columns] = cleaned_df[object_columns].replace(
            to_replace=r'^\s*NA\s*$',
            value='',
            regex=True,
        )
    return cleaned_df


def _append_pending_ra_summary_sheet(
    writer: pd.ExcelWriter,
    detailed_df: pd.DataFrame,
) -> None:
    """Append a teacher-facing worksheet listing pending RAs by student."""
    summary_df = build_pending_ra_summary_dataframe(detailed_df)
    summary_df.to_excel(writer, index=False, sheet_name='Resum')
    ws = writer.sheets['Resum']
    _apply_pending_ra_summary_formatting(ws, len(summary_df))


def build_pending_ra_summary_dataframe(detailed_df: pd.DataFrame) -> pd.DataFrame:
    """Summarize RA rows that still need follow-up into one text block per student."""
    summary_rows: list[dict[str, object]] = []
    columns = detailed_df.columns.tolist()
    for row_number, row in enumerate(detailed_df.to_dict('records'), start=1):
        pending_by_mp: dict[str, list[str]] = {}
        for column_name in columns:
            if column_name == 'estudiant' or not column_name.endswith('RA'):
                continue
            if not _is_pending_ra_grade(row.get(column_name)):
                continue
            mp_code, ra_code = _split_ra_code(column_name)
            pending_by_mp.setdefault(mp_code, []).append(ra_code)

        summary_rows.append(
            {
                '#': row_number,
                'ESTUDIANT': row.get('estudiant', ''),
                'RA PENDENTS': _format_pending_ra_groups(pending_by_mp),
                'OBSERVACIONS': '',
            }
        )

    return pd.DataFrame(summary_rows, columns=['#', 'ESTUDIANT', 'RA PENDENTS', 'OBSERVACIONS'])


def _is_pending_ra_grade(value: object) -> bool:
    """Treat non-passing or unresolved RA grades as pending."""
    if pd.isna(value):
        return False
    if isinstance(value, str):
        normalized = value.strip().upper()
        if normalized in {'PDT', 'EP', 'NA', 'PQ'}:
            return True
        try:
            return float(normalized) < 5
        except ValueError:
            return bool(normalized)
    if isinstance(value, (int, float)):
        return value < 5
    return False


def _split_ra_code(ra_code: str) -> tuple[str, str]:
    """Turn 0647_CF01_2RA into ('0647', '2RA')."""
    parts = ra_code.split('_')
    if len(parts) < 2:
        return ra_code, ra_code
    return parts[0], parts[-1]



def _format_pending_ra_groups(pending_by_mp: dict[str, list[str]]) -> str:
    """Format grouped pending RAs as a multiline checklist per MP code."""
    sections: list[str] = []
    for mp_code in sorted(pending_by_mp):
        ra_codes = sorted(pending_by_mp[mp_code], key=_ra_sort_key)
        lines = "\n".join(f'- {_format_ra_display_code(ra_code)}' for ra_code in ra_codes)
        sections.append(f'{_format_mp_display_name(mp_code)}:\n{lines}')
    return "\n\n".join(sections)


def _ra_sort_key(ra_code: str) -> tuple[int, str]:
    """Sort RA labels numerically so RA10 appears after RA9."""
    digits = ''.join(ch for ch in ra_code if ch.isdigit())
    return int(digits or '0'), ra_code


def _format_ra_display_code(ra_code: str) -> str:
    """Normalize RA codes from forms like 01RA into display labels like RA1."""
    digits = ''.join(ch for ch in ra_code if ch.isdigit())
    if digits:
        return f'RA{int(digits)}'
    return ra_code



def _format_ra_header(header: str) -> str:
    """Break long RA headers over two lines so they stay readable in narrow columns."""
    if not header or header.count('_') < 2:
        return header
    first = header.find('_')
    second = header.find('_', first + 1)
    return f"{header[:second + 1]}\n{header[second + 1:]}"


def _get_mp_for_ra(ra_header: str, mp_codes: list[str]) -> Optional[str]:
    """Map an RA column back to its MP so the correct color can be reused."""
    for mp_code in mp_codes:
        if ra_header.startswith(mp_code + "_") and ra_header.endswith("RA"):
            return mp_code
    return None


def _build_header_lookup(ws: Worksheet) -> dict[str, str]:
    """Map visible header text back to Excel column letters."""
    header_lookup: dict[str, str] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        header_lookup[str(cell.value).replace('\n', '').strip()] = get_column_letter(cell.column)
    return header_lookup


def _apply_row_formatting_to_sheet(
    ws: Worksheet,
    mp_codes_with_em: list[str],
    mp_codes: list[str],
) -> None:
    """Apply the workbook styling directly to an in-memory worksheet."""
    last_row = ws.max_row
    ws.freeze_panes = 'C2'

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center')

    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    type_a_fill = PatternFill(start_color="F1C232", end_color="F1C232", fill_type="solid")
    type_b_fill = PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid")
    alt_fill_even = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    alt_fill_odd = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")

    student_name_width = 40
    mp_column_width = 15
    ra_column_width = 12
    standard_row_height = 25
    mp_codes_with_em_set = set(mp_codes_with_em)

    header_styles: dict[int, tuple[PatternFill, Alignment, Optional[str]]] = {}
    for cell in ws[1]:
        col_letter = get_column_letter(cell.column)
        value = str(cell.value).strip() if cell.value else ""
        formatted_value = None

        if col_letter in {'A', 'B'}:
            fill = gray_fill
            alignment = center_aligned if col_letter == 'A' else left_aligned
        elif value in mp_codes:
            fill = type_a_fill if value in mp_codes_with_em_set else type_b_fill
            alignment = center_aligned
        elif any(value == f"{mp} CENTRE" or value == f"{mp} EMPRESA" for mp in mp_codes_with_em):
            fill = type_a_fill
            alignment = center_aligned
        else:
            mp_code = _get_mp_for_ra(value, mp_codes)
            if mp_code:
                # RA columns inherit their parent MP color so printed workbooks still
                # show which detailed assessments belong to the same module.
                fill = type_a_fill if mp_code in mp_codes_with_em_set else type_b_fill
                alignment = center_aligned
                formatted_value = _format_ra_header(value)
            else:
                fill = PatternFill(fill_type=None)
                alignment = center_aligned

        header_styles[cell.column] = (fill, alignment, formatted_value)

    ws['A1'].value = "#"
    ws['B1'].value = "ESTUDIANT"
    ws.row_dimensions[1].height = 40
    for row_idx in range(2, last_row + 1):
        ws.row_dimensions[row_idx].height = standard_row_height

    for row in ws.iter_rows(min_row=1, max_row=last_row):
        row_idx = row[0].row
        row_fill = alt_fill_even if (row_idx % 2 == 0) else alt_fill_odd
        for cell in row:
            cell.border = border
            if row_idx == 1:
                fill, alignment, formatted_value = header_styles[cell.column]
                cell.font = bold_font
                cell.fill = fill
                cell.alignment = alignment
                if formatted_value is not None:
                    cell.value = formatted_value
            else:
                cell.fill = row_fill
                if cell.column == 1:
                    cell.font = bold_font
                    cell.alignment = center_aligned
                elif cell.column == 2:
                    cell.font = bold_font
                    cell.alignment = left_aligned
                else:
                    cell.alignment = center_aligned

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = student_name_width
    for cell in ws[1][2:]:
        col_letter = get_column_letter(cell.column)
        value = str(cell.value).replace('\n', '').strip() if cell.value else ""
        if value in mp_codes or any(value == f"{mp} CENTRE" or value == f"{mp} EMPRESA" for mp in mp_codes_with_em):
            ws.column_dimensions[col_letter].width = mp_column_width
        else:
            ws.column_dimensions[col_letter].width = ra_column_width

    legend_start_row = last_row + 2
    # The legend stays inside the sheet because exported files are often shared or
    # printed on their own, without extra documentation.
    for offset, (fill, text) in enumerate([
        (type_a_fill, "MP amb estada a l'empresa"),
        (type_b_fill, "MP sense estada a l'empresa")
    ]):
        row = legend_start_row + offset
        ws.row_dimensions[row].height = standard_row_height
        cell1 = ws.cell(row=row, column=1)
        cell2 = ws.cell(row=row, column=2, value=text)
        cell1.fill = fill
        cell2.fill = PatternFill(fill_type=None)
        for cell in (cell1, cell2):
            cell.border = border
            cell.font = bold_font
            cell.alignment = center_aligned if cell.column == 1 else Alignment(wrap_text=True, vertical='center')


def _apply_conditional_formatting_to_sheet(
    ws: Worksheet,
    mp_groups: dict[str, list[str]],
    mp_codes_with_em: list[str],
    mp_codes: list[str],
) -> None:
    """Apply conditional formatting rules directly to an in-memory worksheet."""
    last_student_row = ws.max_row - 3
    if last_student_row < 2:
        return

    red_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    red_font = Font(color="FF0000")
    header_lookup = _build_header_lookup(ws)

    def get_column_for_header(header: str) -> Optional[str]:
        return header_lookup.get(str(header).replace('\n', '').strip())

    def apply_rules_to_column(col_letter: str) -> None:
        range_ref = f'{col_letter}2:{col_letter}{last_student_row}'
        first_cell_ref = f'{col_letter}2'

        # Rules are written against the first row of the range because Excel shifts the
        # relative reference automatically for each cell in the formatted range.
        orange_formula = (
            f'OR(ISNUMBER(SEARCH("PDT",{first_cell_ref})),'
            f'ISNUMBER(SEARCH("EP",{first_cell_ref})),'
            f'ISNUMBER(SEARCH("PQ",{first_cell_ref})))'
        )
        red_fill_formula = (
            f'AND(NOT(ISNUMBER({first_cell_ref})),'
            f'ISERROR(SEARCH("PDT",{first_cell_ref})),'
            f'ISERROR(SEARCH("EP",{first_cell_ref})),'
            f'ISERROR(SEARCH("PQ",{first_cell_ref})))'
        )

        ws.conditional_formatting.add(
            range_ref,
            FormulaRule(formula=[orange_formula], fill=orange_fill, stopIfTrue=True),
        )
        ws.conditional_formatting.add(
            range_ref,
            FormulaRule(formula=[red_fill_formula], fill=red_fill, stopIfTrue=True),
        )
        ws.conditional_formatting.add(
            range_ref,
            FormulaRule(formula=[f'AND(ISNUMBER({first_cell_ref}), {first_cell_ref}<5)'], font=red_font, stopIfTrue=True),
        )

    for mp_code in mp_codes:
        for ra_code in mp_groups.get(mp_code, []):
            col_letter = get_column_for_header(ra_code)
            if col_letter:
                apply_rules_to_column(col_letter)

    for mp_code in mp_codes:
        related_headers = [mp_code]
        if mp_code in mp_codes_with_em:
            related_headers.insert(0, f"{mp_code} EMPRESA")
            related_headers.insert(0, f"{mp_code} CENTRE")
        for header in related_headers:
            col_letter = get_column_for_header(header)
            if col_letter:
                apply_rules_to_column(col_letter)



def _apply_pending_ra_summary_formatting(ws: Worksheet, row_count: int) -> None:
    """Apply readable formatting to the pending-RA summary sheet."""
    ws.freeze_panes = 'C2'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)
    header_font = Font(bold=True, color='FFFFFF')
    shared_header_font = Font(bold=True)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='1C4587', end_color='1C4587', fill_type='solid')
    shared_header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    alt_fill_even = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    alt_fill_odd = PatternFill(start_color='EEF6FF', end_color='EEF6FF', fill_type='solid')

    ws.row_dimensions[1].height = 34
    for row_idx in range(2, row_count + 2):
        row_values = [ws.cell(row=row_idx, column=column).value for column in range(1, 5)]
        max_lines = max(_count_display_lines(value) for value in row_values)
        ws.row_dimensions[row_idx].height = max(24, 16 * max_lines + 8)

    for row in ws.iter_rows(min_row=1, max_row=row_count + 1, max_col=4):
        row_idx = row[0].row
        row_fill = alt_fill_even if row_idx % 2 == 0 else alt_fill_odd
        for cell in row:
            cell.border = border
            if row_idx == 1:
                if cell.column in {1, 2}:
                    cell.font = shared_header_font
                    cell.fill = shared_header_fill
                else:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.alignment = center_aligned if cell.column != 2 else left_aligned
            else:
                cell.fill = row_fill
                if cell.column == 1:
                    cell.font = bold_font
                    cell.alignment = center_aligned
                else:
                    cell.alignment = left_aligned

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 44
    ws.column_dimensions['D'].width = 28



def _count_display_lines(value: object) -> int:
    """Estimate wrapped spreadsheet lines for row-height sizing."""
    if value is None:
        return 1
    text = str(value)
    if not text:
        return 1
    return max(1, text.count("\n") + 1)
