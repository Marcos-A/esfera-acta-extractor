import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import re
import os
import time

def _extract_mp_codes_from_columns(columns: list[str]) -> list[str]:
    """Extracts unique 4-digit MP codes from a list of column names."""
    mp_codes = set()
    # Matches a 4-digit MP code at the start of a column name,
    # optionally followed by specific suffixes like _...RA, CENTRE, or EMPRESA.
    # This also matches plain 4-digit MP codes.
    mp_pattern = re.compile(r"^(?P<code>[A-Za-z0-9]{3,5})(?:_.*RA| CENTRE| EMPRESA)?$")
    for col in columns:
        match = mp_pattern.match(col)
        if match:
            # Ensure 'code' group exists before accessing, though pattern implies it should
            code_match_group = match.group('code') 
            if code_match_group:
                code = code_match_group
                # Explicitly exclude 'estudiant' (case-insensitive) from being an MP code
                if code.lower() != 'estudiant':
                    mp_codes.add(code)
    return sorted(list(mp_codes))

# Constants for file reading retries
MAX_READ_ATTEMPTS = 3
READ_RETRY_DELAY_SECONDS = 5

def generate_summary_report(source_xlsx_path: str, output_summary_path: str, include_weighting: bool):
    """Generates a summary Excel report from a detailed grade Excel file."""
    df_source = None
    
    # 1. Read Source XLSX with retries for temporary issues
    for attempt in range(1, MAX_READ_ATTEMPTS + 1):
        try:
            footer_rows = 1 if include_weighting else 0
            df_source = pd.read_excel(source_xlsx_path, sheet_name=0, skipfooter=footer_rows)
            break  # Success, exit retry loop
        except FileNotFoundError:
            print(f"[ERROR summary_generator] Source file not found: {source_xlsx_path}. This usually means it was deleted after being listed.")
            return # Definite error, no retry needed for this specific case
        except Exception as e: # Catch other errors that might be temporary (e.g., file lock, incomplete write)
            print(f"[WARN summary_generator] Attempt {attempt}/{MAX_READ_ATTEMPTS} failed to read '{os.path.basename(source_xlsx_path)}': {type(e).__name__} - {e}")
            if attempt < MAX_READ_ATTEMPTS:
                print(f"Retrying in {READ_RETRY_DELAY_SECONDS} seconds...")
                time.sleep(READ_RETRY_DELAY_SECONDS)
            else:
                print(f"[ERROR summary_generator] All {MAX_READ_ATTEMPTS} attempts to read '{os.path.basename(source_xlsx_path)}' failed. Skipping summary for this file.")
                return # All retries failed

    if df_source is None:
        # This should ideally be caught by a return in the loop, but acts as a final safeguard.
        print(f"[ERROR summary_generator] Could not load data from {source_xlsx_path} after all attempts. Skipping.")
        return

    all_source_columns = df_source.columns.tolist()

    # 2. Identify MP Codes
    mp_codes = _extract_mp_codes_from_columns(all_source_columns)

    mp_info_for_summary = []  # Stores dicts of {'mp_code', 'col_name', 'type_a', 'header_bg'}

    # 3. Determine columns to include for each MP
    for mp_code in mp_codes:
        # Determine type for summary and column to include
        # Type A for summary: has a '{MP_CODE} CENTRE' column. We take this column.
        # Type B for summary: does not have CENTRE. We take the plain '{MP_CODE}' column.
        centre_col_name = f"{mp_code} CENTRE"
        is_type_a_for_summary = centre_col_name in all_source_columns

        if is_type_a_for_summary:
            mp_info_for_summary.append({'mp_code': mp_code, 'col_name': centre_col_name, 'type_a': True, 'header_bg': "F1C232"})
        else:  # Type B for summary
            plain_mp_col_name = mp_code
            if plain_mp_col_name in all_source_columns:
                mp_info_for_summary.append({'mp_code': mp_code, 'col_name': plain_mp_col_name, 'type_a': False, 'header_bg': "B4A7D6"})
            else:
                print(f"[INFO summary_generator] MP {mp_code} is Type B for summary, but its plain column '{plain_mp_col_name}' not found. Skipping.")
                continue

    # 4. Construct Summary DataFrame
    if not mp_info_for_summary:
        print(f"[INFO summary_generator] No MPs to include in summary for {source_xlsx_path}. No summary file will be generated.")
        return

    # Find the actual student column name (case-insensitive)
    student_col_actual_name = None
    for col in df_source.columns:
        if col.lower() == 'estudiant':
            student_col_actual_name = col
            break
    
    if not student_col_actual_name:
        print(f"[ERROR summary_generator] Could not find a student column (e.g., 'estudiant') in {source_xlsx_path}. Cannot generate summary.")
        return

    summary_df_cols_to_select = [student_col_actual_name] + [info['col_name'] for info in mp_info_for_summary]
    
    # Ensure all selected columns exist in df_source (should be guaranteed by checks above)
    summary_df = df_source[summary_df_cols_to_select].copy()
    # Rename student column to 'estudiant' for consistency in the summary file if it was different
    if student_col_actual_name != 'estudiant':
        summary_df.rename(columns={student_col_actual_name: 'estudiant'}, inplace=True)
    
    # Add numbered column as the first column
    summary_df.insert(0, '#', range(1, len(summary_df) + 1))

    # 5. Write to New XLSX and Apply Formatting
    try:
        with pd.ExcelWriter(output_summary_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, index=False, sheet_name='Summary')

        wb = load_workbook(output_summary_path)
        ws = wb['Summary']

        ws.freeze_panes = 'C2'  # Freeze first two rows and first column

        thin_border_side = Side(border_style="thin", color="000000")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # Format header row
        for col_idx, column_title_from_df in enumerate(summary_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.border = thin_border
            
            if column_title_from_df == '#':
                cell.value = '#'
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif column_title_from_df.lower() == 'estudiant':
                cell.value = "ESTUDIANT" # Capitalize header for display
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True) # Left-align header
            else:
                # For MP columns
                cell.value = column_title_from_df # Use the original MP column name for the header
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                current_mp_info = next((info for info in mp_info_for_summary if info['col_name'] == column_title_from_df), None)
                if current_mp_info:
                    cell.fill = PatternFill(start_color=current_mp_info['header_bg'], end_color=current_mp_info['header_bg'], fill_type="solid")

        # First, collect all data validations to apply to columns
        data_validations = {}
        for col_idx, col_name in enumerate(summary_df.columns, 1):
            if col_name.lower() == 'estudiant':
                continue  # Skip student column
                
            current_mp_info = next((info for info in mp_info_for_summary if info['col_name'] == col_name), None)
            if current_mp_info and current_mp_info['type_a']:
                # If this is a CENTRE column, restrict to 0-9
                if current_mp_info['col_name'].endswith('CENTRE'):
                    dv = DataValidation(
                        type="decimal",
                        operator="between",
                        formula1=0,
                        formula2=9,
                        allow_blank=True,
                        errorTitle='Valor invàlid',
                        error='Introdueix un número entre 0 i 9 amb un màxim de dos decimals.',
                        showErrorMessage=True
                    )
                else:
                    # For other Type A columns (e.g., MP final grade), allow 0-10
                    dv = DataValidation(
                        type="decimal",
                        operator="between",
                        formula1=0,
                        formula2=10,
                        allow_blank=True,
                        errorTitle='Valor invàlid',
                        error='Introdueix un número entre 0 i 10 amb un màxim de dos decimals.',
                        showErrorMessage=True
                    )
                # Store the validation with column index as key
                data_validations[col_idx] = dv
                ws.add_data_validation(dv)
            elif current_mp_info and not current_mp_info['type_a']:  # Type B columns
                # For Type B columns, allow only integers 0-10
                dv = DataValidation(
                    type="whole",
                    operator="between",
                    formula1=0,
                    formula2=10,
                    allow_blank=True,
                    errorTitle='Valor invàlid',
                    error='Introdueix un número enter entre 0 i 10.',
                    showErrorMessage=True
                )
                data_validations[col_idx] = dv
                ws.add_data_validation(dv)
        
        # Define alternating row colors for data rows
        light_green = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        dark_green = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")

        # First, apply all existing formatting
        for row_idx, row in enumerate(ws.iter_rows(min_row=1), 1):
            is_data_row = 1 < row_idx <= len(summary_df) + 1  # Only actual data rows
            
            # Determine row fill for data rows
            if is_data_row:
                row_fill = light_green if (row_idx % 2 == 0) else dark_green
            
            for col_idx, cell in enumerate(row, 1):
                col_name = summary_df.columns[col_idx - 1] if col_idx <= len(summary_df.columns) else ""
                
                # Apply borders to all cells
                cell.border = thin_border
                
                # Apply cell-specific formatting
                if col_name == '#':  # Number column
                    if row_idx == 1:  # Header row
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:  # Number cells
                        cell.fill = row_fill if is_data_row else PatternFill(start_color="D9F7F0", end_color="D9F7F0", fill_type="solid")
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name.lower() == 'estudiant':  # Student column
                    if row_idx == 1:  # Header row
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        cell.font = Font(bold=True)
                    else:  # Student names
                        cell.fill = row_fill if is_data_row else PatternFill(start_color="D9F7F0", end_color="D9F7F0", fill_type="solid")
                        cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif is_data_row and col_name:  # Data cells in data rows
                    cell.fill = row_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:  # MP columns and other cells
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply data validation if it exists for this column
                if col_idx in data_validations:
                    col_letter = get_column_letter(col_idx)
                    data_validations[col_idx].add(f"{col_letter}{row_idx}")
        
        # For Type B MP data cells, no specific data validation beyond default cell properties (center alignment)
        
        # Apply number formatting based on MP type
        for col_idx, col_name in enumerate(summary_df.columns, 1):
            # Skip non-MP columns
            if col_name in ['#', 'estudiant']:
                continue
                
            # Find the MP info for this column
            mp_info = next((mp for mp in mp_info_for_summary if mp['col_name'] == col_name), None)
            if not mp_info:
                continue
                
            # Format the column based on MP type
            for row in ws.iter_rows(min_row=2, max_row=len(summary_df) + 1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if mp_info['type_a']:
                        cell.number_format = '0.00'  # 2 decimal places for Type A
                    else:
                        cell.number_format = '0'  # Integer for Type B
        
        # Add conditional formatting to highlight empty or non-numeric cells in red
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        for col_idx, col_name in enumerate(summary_df.columns, 1):
            if col_name.lower() == 'estudiant':
                continue  # Skip student column
                
            # Create a formula that checks if the cell is empty or not a number
            col_letter = get_column_letter(col_idx)
            range_str = f"{col_letter}2:{col_letter}{ws.max_row}"
            
            # Add conditional formatting rule for empty cells
            ws.conditional_formatting.add(
                range_str,
                FormulaRule(
                    formula=[f'OR(ISBLANK({col_letter}2), NOT(ISNUMBER({col_letter}2)))'],
                    stopIfTrue=True,
                    fill=red_fill
                )
            )

        # Auto-adjust column widths
        for col_idx, column_name in enumerate(summary_df.columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            if column_name:
                max_length = max(max_length, len(str(column_name)))
            for i in range(1, ws.max_row + 1):
                if ws.cell(row=i, column=col_idx).value is not None:
                    max_length = max(max_length, len(str(ws.cell(row=i, column=col_idx).value)))
            adjusted_width = (max_length + 2) if max_length > 0 else len(str(column_name)) + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add legend at the bottom of the sheet
        legend_start_row = ws.max_row + 2  # Leave one empty row after the data
        
        # Create a bold font for the legend
        bold_font = Font(bold=True)
        
        # First legend row (Type A MP)
        ws.cell(row=legend_start_row, column=1).fill = PatternFill(start_color="F1C232", end_color="F1C232", fill_type="solid")
        cell_a1 = ws.cell(row=legend_start_row, column=2, value="MP amb estada a l'empresa\nNOTA PONDERADA AL 90% amb 2 decimals")
        cell_a1.font = bold_font
        cell_a1.alignment = Alignment(wrap_text=True, vertical='center')
        
        # Second legend row (Type B MP)
        ws.cell(row=legend_start_row + 1, column=1).fill = PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid")
        cell_b1 = ws.cell(row=legend_start_row + 1, column=2, value="MP sense estada a l'empresa\nNOTA SOBRE 10 i sense decimals")
        cell_b1.font = bold_font
        cell_b1.alignment = Alignment(wrap_text=True, vertical='center')
        
        # Set row heights and column widths for the legend
        ws.row_dimensions[legend_start_row].height = 40
        ws.row_dimensions[legend_start_row + 1].height = 40
        
        # Adjust column widths to fit content
        ws.column_dimensions['A'].width = 5  # First column (color indicator)
        ws.column_dimensions['B'].width = 35  # Second column (text)
        
        # Apply borders to legend cells
        for row in range(legend_start_row, legend_start_row + 2):
            for col in [1, 2]:
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.font = bold_font  # Ensure all legend text is bold

        wb.save(output_summary_path)

    except Exception as e:
        # Clean up partially created file if error occurs during writing/formatting
        if os.path.exists(output_summary_path):
            try:
                os.remove(output_summary_path)
            except Exception as remove_e:
                print(f"[ERROR summary_generator] Could not remove partial file {output_summary_path}: {remove_e}")
        print(f"[ERROR summary_generator] Error writing or formatting summary file {output_summary_path}: {e}")

