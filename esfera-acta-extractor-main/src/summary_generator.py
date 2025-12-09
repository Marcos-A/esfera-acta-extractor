import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import re
import os
import time

def _extract_mp_codes_from_columns(columns: list[str]) -> list[str]:
    """Extracts unique 4-digit MP codes from a list of column names."""
    mp_codes = set()
    mp_pattern = re.compile(r"^(?P<code>[A-Za-z0-9]{3,5})(?:_.*RA| CENTRE| EMPRESA)?$")
    for col in columns:
        match = mp_pattern.match(col)
        if match:
            code = match.group('code')
            if code and code.lower() != 'estudiant':
                mp_codes.add(code)
    return sorted(list(mp_codes))

MAX_READ_ATTEMPTS = 3
READ_RETRY_DELAY_SECONDS = 5

def generate_summary_report(source_xlsx_path: str, output_summary_path: str):
    df_source = None

    for attempt in range(1, MAX_READ_ATTEMPTS + 1):
        try:
            df_source = pd.read_excel(source_xlsx_path, sheet_name=0)
            break
        except FileNotFoundError:
            print(f"[ERROR summary_generator] Source file not found: {source_xlsx_path}.")
            return
        except Exception as e:
            print(f"[WARN summary_generator] Attempt {attempt}/{MAX_READ_ATTEMPTS} failed to read '{os.path.basename(source_xlsx_path)}': {type(e).__name__} - {e}")
            if attempt < MAX_READ_ATTEMPTS:
                print(f"Retrying in {READ_RETRY_DELAY_SECONDS} seconds...")
                time.sleep(READ_RETRY_DELAY_SECONDS)
            else:
                print(f"[ERROR summary_generator] All {MAX_READ_ATTEMPTS} attempts failed. Skipping summary.")
                return

    if df_source is None:
        print(f"[ERROR summary_generator] Could not load data from {source_xlsx_path}. Skipping.")
        return

    # Detect actual student data rows by finding the first completely empty row
    empty_row_index = df_source[df_source.isnull().all(axis=1)].index
    if not empty_row_index.empty:
        df_source = df_source.loc[:empty_row_index[0] - 1]

    all_source_columns = df_source.columns.tolist()
    mp_codes = _extract_mp_codes_from_columns(all_source_columns)
    mp_info_for_summary = []

    for mp_code in mp_codes:
        centre_col = f"{mp_code} CENTRE"
        if centre_col in all_source_columns:
            mp_info_for_summary.append({'mp_code': mp_code, 'col_name': centre_col, 'type_a': True, 'header_bg': "F1C232"})
        else:
            plain_mp = mp_code
            if plain_mp in all_source_columns:
                mp_info_for_summary.append({'mp_code': mp_code, 'col_name': plain_mp, 'type_a': False, 'header_bg': "B4A7D6"})

    if not mp_info_for_summary:
        print(f"[INFO summary_generator] No MPs to include in summary for {source_xlsx_path}.")
        return

    student_col_actual_name = next((col for col in df_source.columns if col.lower() == 'estudiant'), None)
    if not student_col_actual_name:
        print(f"[ERROR summary_generator] Could not find 'estudiant' column in {source_xlsx_path}.")
        return

    selected_cols = [student_col_actual_name] + [info['col_name'] for info in mp_info_for_summary]
    summary_df = df_source[selected_cols].copy()
    if student_col_actual_name != 'estudiant':
        summary_df.rename(columns={student_col_actual_name: 'estudiant'}, inplace=True)

    summary_df.insert(0, '#', range(1, len(summary_df) + 1))
    with pd.ExcelWriter(output_summary_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, index=False, sheet_name='Summary')

    # print(f"[INFO summary_generator] Summary report created: {output_summary_path}")

    # Find the actual student column name (case-insensitive)
    student_col_actual_name = None
    for col in df_source.columns:
        if col.lower() == 'estudiant':
            student_col_actual_name = col
            break
    
    if not student_col_actual_name:
        print(f"[ERROR summary_generator] Could not find a student column (e.g., 'estudiant') in {source_xlsx_path}.")
        return

    summary_df_cols_to_select = [student_col_actual_name] + [info['col_name'] for info in mp_info_for_summary]
    
    # Ensure all selected columns exist in df_source (should be guaranteed by checks above)
    summary_df = df_source[summary_df_cols_to_select].copy()
    # Rename student column to 'estudiant' for consistency in the summary file if it was different
    if student_col_actual_name != 'estudiant':
        summary_df.rename(columns={student_col_actual_name: 'estudiant'}, inplace=True)
    
    # Add numbered column as the first column
    summary_df.insert(0, '#', range(1, len(summary_df) + 1))

    # 5. Write to New XLSX and Apply Formatting
    try:
        with pd.ExcelWriter(output_summary_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, index=False, sheet_name='Summary')

        wb = load_workbook(output_summary_path)
        ws = wb['Summary']

        ws.freeze_panes = 'C2'  # Freeze first two rows and first column

        thin_border_side = Side(border_style="thin", color="000000")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # Define standard row height
        STANDARD_ROW_HEIGHT = 25 # Define the standard row height here

        # Set row heights for all rows
        ws.row_dimensions[1].height = 40  # Header row height
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = STANDARD_ROW_HEIGHT

        # Format header row
        for col_idx, column_title_from_df in enumerate(summary_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.border = thin_border
            
            if column_title_from_df == '#':
                cell.value = '#'
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif column_title_from_df.lower() == 'estudiant':
                cell.value = "ESTUDIANT" # Capitalize header for display
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True) # Left-align header
            else:
                # For MP columns
                cell.value = column_title_from_df # Use the original MP column name for the header
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                current_mp_info = next((info for info in mp_info_for_summary if info['col_name'] == column_title_from_df), None)
                if current_mp_info:
                    cell.fill = PatternFill(start_color=current_mp_info['header_bg'], end_color=current_mp_info['header_bg'], fill_type="solid")

        # First, collect all data validations to apply to columns
        data_validations = {}
        for col_idx, col_name in enumerate(summary_df.columns, 1):
            if col_name.lower() == 'estudiant':
                continue  # Skip student column
                
            current_mp_info = next((info for info in mp_info_for_summary if info['col_name'] == col_name), None)
            if current_mp_info and current_mp_info['type_a']:
                # If this is a CENTRE column, restrict to 0-9
                if current_mp_info['col_name'].endswith('CENTRE'):
                    dv = DataValidation(
                        type="decimal",
                        operator="between",
                        formula1=0,
                        formula2=9,
                        allow_blank=True,
                        errorTitle='Valor invàlid',
                        error='Introdueix un número entre 0 i 9 amb un màxim de dos decimals.',
                        showErrorMessage=True
                    )
                else:
                    # For other Type A columns (e.g., MP final grade), allow 0-10
                    dv = DataValidation(
                        type="decimal",
                        operator="between",
                        formula1=0,
                        formula2=10,
                        allow_blank=True,
                        errorTitle='Valor invàlid',
                        error='Introdueix un número entre 0 i 10 amb un màxim de dos decimals.',
                        showErrorMessage=True
                    )
                # Store the validation with column index as key
                data_validations[col_idx] = dv
                ws.add_data_validation(dv)
            elif current_mp_info and not current_mp_info['type_a']:  # Type B columns
                # For Type B columns, allow only integers 0-10
                dv = DataValidation(
                    type="whole",
                    operator="between",
                    formula1=0,
                    formula2=10,
                    allow_blank=True,
                    errorTitle='Valor invàlid',
                    error='Introdueix un número enter entre 0 i 10.',
                    showErrorMessage=True
                )
                data_validations[col_idx] = dv
                ws.add_data_validation(dv)
        
        # Define alternating row colors for data rows
        light_green = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        dark_green = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")

        # First, apply all existing formatting
        for row_idx, row in enumerate(ws.iter_rows(min_row=1), 1):
            is_data_row = 1 < row_idx <= len(summary_df) + 1  # Only actual data rows
            
            # Determine row fill for data rows
            if is_data_row:
                row_fill = light_green if (row_idx % 2 == 0) else dark_green
            
            for col_idx, cell in enumerate(row, 1):
                col_name = summary_df.columns[col_idx - 1] if col_idx <= len(summary_df.columns) else ""
                
                # Apply borders to all cells
                cell.border = thin_border
                
                # Apply cell-specific formatting
                if col_name == '#':  # Number column
                    if row_idx == 1:  # Header row
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:  # Number cells
                        cell.fill = row_fill if is_data_row else PatternFill(start_color="D9F7F0", end_color="D9F7F0", fill_type="solid")
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name.lower() == 'estudiant':  # Student column
                    if row_idx == 1:  # Header row
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        cell.font = Font(bold=True)
                    else:  # Student names
                        cell.fill = row_fill if is_data_row else PatternFill(start_color="D9F7F0", end_color="D9F7F0", fill_type="solid")
                        cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif is_data_row and col_name:  # Data cells in data rows
                    cell.fill = row_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:  # MP columns and other cells
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply data validation if it exists for this column
                if col_idx in data_validations:
                    col_letter = get_column_letter(col_idx)
                    data_validations[col_idx].add(f"{col_letter}{row_idx}")
        
        # For Type B MP data cells, no specific data validation beyond default cell properties (center alignment)
        
        # Apply number formatting based on MP type
        for col_idx, col_name in enumerate(summary_df.columns, 1):
            # Skip non-MP columns
            if col_name in ['#', 'estudiant']:
                continue
                
            # Find the MP info for this column
            mp_info = next((mp for mp in mp_info_for_summary if mp['col_name'] == col_name), None)
            if not mp_info:
                continue
                
            # Format the column based on MP type
            for row in ws.iter_rows(min_row=2, max_row=len(summary_df) + 1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if mp_info['type_a']:
                        cell.number_format = '0.00'  # 2 decimal places for Type A
                    else:
                        cell.number_format = '0'  # Integer for Type B
        
        # Add conditional formatting to highlight empty or non-numeric cells in red
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # New fill for specific strings
        orange_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")

        # Define red font for numbers < 5
        red_font = Font(color="FF0000")

        for col_idx, col_name in enumerate(summary_df.columns, 1):
            # IMPORTANT: Skip the first two columns ('#' and 'estudiant') for conditional formatting
            if col_name.lower() == 'estudiant' or col_name == '#':
                continue
                
            # Determine the threshold for red text based on column name
            threshold = 4.5 if col_name.endswith('CENTRE') else 5

            # Create a formula that checks if the cell is empty or not a number
            col_letter = get_column_letter(col_idx)
            range_str = f"{col_letter}2:{col_letter}{ws.max_row}"
            
            # Add conditional formatting rule for cells containing specific strings (PDT, EP, NA, PQ)
            # This rule should be applied *before* the general non-numeric/empty rule
            # because conditional formatting rules are applied in the order they are added.
            # The stopIfTrue=True ensures that if this rule applies, the next rule is not checked.
            ws.conditional_formatting.add(
                range_str,
                FormulaRule(
                    formula=[
                        f'OR(UPPER({col_letter}2)="PDT", '
                        f'UPPER({col_letter}2)="EP", '
                        f'UPPER({col_letter}2)="NA", '
                        f'UPPER({col_letter}2)="PQ")'
                    ],
                    stopIfTrue=True,
                    fill=orange_fill
                )
            )

            # Add conditional formatting rule for empty cells or non-numeric values
            ws.conditional_formatting.add(
                range_str,
                FormulaRule(
                    formula=[f'OR(ISBLANK({col_letter}2), NOT(ISNUMBER({col_letter}2)))'],
                    stopIfTrue=True,
                    fill=red_fill
                )
            )

            # Add conditional formatting rule for numerical values smaller than the determined threshold
            ws.conditional_formatting.add(
                range_str,
                FormulaRule(
                    formula=[f'AND(ISNUMBER({col_letter}2), {col_letter}2<{threshold})'],
                    font=red_font
                )
            )

        # Auto-adjust column widths
        for col_idx, column_name in enumerate(summary_df.columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            if column_name:
                max_length = max(max_length, len(str(column_name)))
            for i in range(1, ws.max_row + 1):
                if ws.cell(row=i, column=col_idx).value is not None:
                    max_length = max(max_length, len(str(ws.cell(row=i, column=col_idx).value)))
            adjusted_width = (max_length + 2) if max_length > 0 else len(str(column_name)) + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add legend at the bottom of the sheet
        legend_start_row = ws.max_row + 2  # Leave one empty row after the data

        # Create a bold font and border for the legend
        bold_font = Font(bold=True)
        centered_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        legend_entries = [
            ("F1C232", "MP amb estada a l'empresa\nNOTA PONDERADA AL 90% amb 2 decimals"),
            ("B4A7D6", "MP sense estada a l'empresa\nNOTA SOBRE 10 i sense decimals"),
        ]

        for i, (color, text) in enumerate(legend_entries):
            row = legend_start_row + i
            color_cell = ws.cell(row=row, column=1)
            label_cell = ws.cell(row=row, column=2, value=text)

            # Apply styles
            color_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            color_cell.border = thin_border
            color_cell.alignment = centered_wrap

            label_cell.fill = PatternFill(fill_type=None)  # No background color
            label_cell.font = bold_font
            label_cell.border = thin_border
            label_cell.alignment = left_wrap

            # Set row height for visibility
            ws.row_dimensions[row].height = 45

        # Adjust column widths for legend
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50


        # Apply borders to legend cells
        for row in range(legend_start_row, legend_start_row + 2):
            for col in [1, 2]:
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.font = bold_font  # Ensure all legend text is bold

        wb.save(output_summary_path)

    except Exception as e:
        # Clean up partially created file if error occurs during writing/formatting
        if os.path.exists(output_summary_path):
            try:
                os.remove(output_summary_path)
            except Exception as remove_e:
                print(f"[ERROR summary_generator] Could not remove partial file {output_summary_path}: {remove_e}")
        print(f"[ERROR summary_generator] Error writing or formatting summary file {output_summary_path}: {e}")
