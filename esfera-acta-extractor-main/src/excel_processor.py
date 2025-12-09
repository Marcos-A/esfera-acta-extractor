"""
Excel processing module for generating and formatting grade reports.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from typing import Optional


def apply_row_formatting(
    workbook_path: str,
    mp_codes_with_em: list[str],
    mp_codes: list[str],
) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from typing import Optional

    wb = load_workbook(workbook_path)
    ws = wb.active
    last_row = ws.max_row
    ws.freeze_panes = 'C2'

    # Styles
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center')

    # Fill colors
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    type_a_fill = PatternFill(start_color="F1C232", end_color="F1C232", fill_type="solid")
    type_b_fill = PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid")
    alt_fill_even = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    alt_fill_odd = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")

    # Dimensions
    STUDENT_NAME_WIDTH = 40
    MP_COLUMN_WIDTH = 15
    RA_COLUMN_WIDTH = 12
    STANDARD_ROW_HEIGHT = 25

    def format_ra_header(header: str) -> str:
        if not header or header.count('_') < 2:
            return header
        first = header.find('_')
        second = header.find('_', first + 1)
        return f"{header[:second + 1]}\n{header[second + 1:]}"

    def get_mp_for_ra(ra_header: str) -> Optional[str]:
        for mp_code in mp_codes:
            if ra_header.startswith(mp_code + "_") and ra_header.endswith("RA"):
                return mp_code
        return None

    # Adjust header row
    ws['A1'].value = "#"
    ws['B1'].value = "ESTUDIANT"

    ws.row_dimensions[1].height = 40
    for row_idx in range(2, last_row + 1):
        ws.row_dimensions[row_idx].height = STANDARD_ROW_HEIGHT

    for row in ws.iter_rows(min_row=1, max_row=last_row):
        row_idx = row[0].row
        for cell in row:
            col_letter = get_column_letter(cell.column)
            val = str(cell.value).strip() if cell.value else ""
            cell.border = border

            if row_idx == 1:
                cell.font = bold_font
                if col_letter == 'A' or col_letter == 'B':
                    cell.fill = gray_fill
                    cell.alignment = center_aligned if col_letter == 'A' else left_aligned
                elif val in mp_codes:
                    cell.fill = type_a_fill if val in mp_codes_with_em else type_b_fill
                    cell.alignment = center_aligned
                elif any(val == f"{mp} CENTRE" or val == f"{mp} EMPRESA" for mp in mp_codes_with_em):
                    cell.fill = type_a_fill
                    cell.alignment = center_aligned
                elif get_mp_for_ra(val):
                    mp_code = get_mp_for_ra(val)
                    if mp_code:
                        cell.fill = type_a_fill if mp_code in mp_codes_with_em else type_b_fill
                        cell.value = format_ra_header(val)
                        cell.alignment = center_aligned
                else:
                    cell.alignment = center_aligned
            else:
                cell.fill = alt_fill_even if (row_idx % 2 == 0) else alt_fill_odd
                if col_letter == 'A':
                    cell.font = bold_font
                    cell.alignment = center_aligned
                elif col_letter == 'B':
                    cell.font = bold_font
                    cell.alignment = left_aligned
                else:
                    cell.alignment = center_aligned

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = STUDENT_NAME_WIDTH
    for cell in ws[1][2:]:
        col_letter = get_column_letter(cell.column)
        val = str(cell.value).strip() if cell.value else ""
        if val in mp_codes:
            ws.column_dimensions[col_letter].width = MP_COLUMN_WIDTH
        elif any(val == f"{mp} CENTRE" or val == f"{mp} EMPRESA" for mp in mp_codes_with_em):
            ws.column_dimensions[col_letter].width = MP_COLUMN_WIDTH
        elif get_mp_for_ra(val):
            ws.column_dimensions[col_letter].width = RA_COLUMN_WIDTH
        else:
            ws.column_dimensions[col_letter].width = RA_COLUMN_WIDTH

    # Insert legend after student list with one empty row in between
    legend_start_row = last_row + 2

    for offset, (fill, text) in enumerate([
        (type_a_fill, "MP amb estada a l'empresa"),
        (type_b_fill, "MP sense estada a l'empresa")
    ]):
        row = legend_start_row + offset
        ws.row_dimensions[row].height = STANDARD_ROW_HEIGHT
        cell1 = ws.cell(row=row, column=1)
        cell2 = ws.cell(row=row, column=2, value=text)
        
        cell1.fill = fill
        cell2.fill = PatternFill(fill_type=None)  # No fill
        for cell in (cell1, cell2):
            cell.border = border
            cell.font = bold_font
            cell.alignment = center_aligned if cell.column == 1 else Alignment(wrap_text=True, vertical='center')

    wb.save(workbook_path)


def apply_conditional_formatting(
    workbook_path: str,
    mp_groups: dict[str, list[str]],
    mp_codes_with_em: list[str],
    mp_codes: list[str],
    preserve_na: bool = True
) -> None:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    wb = load_workbook(workbook_path)
    ws = wb.active
    last_student_row = ws.max_row - 3  # 1 empty row + 2 legend rows

    red_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    red_font = Font(color="FF0000")

    def get_column_for_header(header: str) -> str:
        header_clean = str(header).replace('\n', '')
        for cell in ws[1]:
            if cell.value and str(cell.value).replace('\n', '') == header_clean:
                return get_column_letter(cell.column)
        return None

    def apply_rules_to_column(col_letter: str):
        for row in range(2, last_student_row + 1):
            cell_ref = f'{col_letter}{row}'

            # Orange fill: PDT, EP, PQ (and NA only when not preserving NA)
            if preserve_na:
                orange_formula = (
                    f'OR(ISNUMBER(SEARCH("PDT",{cell_ref})),'
                    f'ISNUMBER(SEARCH("EP",{cell_ref})),'
                    f'ISNUMBER(SEARCH("PQ",{cell_ref})))'
                )
            else:
                orange_formula = (
                    f'OR(ISNUMBER(SEARCH("PDT",{cell_ref})),'
                    f'ISNUMBER(SEARCH("EP",{cell_ref})),'
                    f'ISNUMBER(SEARCH("NA",{cell_ref})),'
                    f'ISNUMBER(SEARCH("PQ",{cell_ref})))'
                )

            orange_rule = FormulaRule(
                formula=[orange_formula],
                fill=orange_fill,
                stopIfTrue=True
            )
            ws.conditional_formatting.add(cell_ref, orange_rule)

            # Red fill for invalid entries (non-number and not one of the above)
            red_fill_formula = (
                f'AND(NOT(ISNUMBER({cell_ref})),'
                f'ISERROR(SEARCH("PDT",{cell_ref})),'
                f'ISERROR(SEARCH("EP",{cell_ref})),'
                f'ISERROR(SEARCH("PQ",{cell_ref})))'
            )

            red_fill_rule = FormulaRule(
                formula=[red_fill_formula],
                fill=red_fill,
                stopIfTrue=True
            )
            ws.conditional_formatting.add(cell_ref, red_fill_rule)

            # If preserving NA, add an explicit rule to mark literal "NA" as invalid
            if preserve_na:
                na_red_rule = FormulaRule(
                    formula=[f'UPPER({cell_ref})="NA"'],
                    fill=red_fill,
                    stopIfTrue=True
                )
                ws.conditional_formatting.add(cell_ref, na_red_rule)

            # Red font if number < 5
            red_font_rule = FormulaRule(
                formula=[f'AND(ISNUMBER({cell_ref}), {cell_ref}<5)'],
                font=red_font,
                stopIfTrue=True
            )
            ws.conditional_formatting.add(cell_ref, red_font_rule)

    # Apply to RA columns
    for mp_code in mp_codes:
        for ra_code in mp_groups.get(mp_code, []):
            col_letter = get_column_for_header(ra_code)
            if col_letter:
                apply_rules_to_column(col_letter)

    # Apply to MP-related columns (CENTRE, EMPRESA, MP)
    for mp_code in mp_codes:
        related_headers = [mp_code]
        if mp_code in mp_codes_with_em:
            related_headers.insert(0, f"{mp_code} EMPRESA")
            related_headers.insert(0, f"{mp_code} CENTRE")
        for header in related_headers:
            col_letter = get_column_for_header(header)
            if col_letter:
                apply_rules_to_column(col_letter)

    wb.save(workbook_path)


def export_excel_with_spacing(
    df: pd.DataFrame,
    output_path: str,
    mp_codes_with_em: list[str],
    mp_codes: list[str],
    preserve_na: bool = True,
) -> None:
    """
    Export DataFrame to Excel with specific column spacing after each MP's RAs.
    Adds a "#" column with sequential numbers at the beginning.

    preserve_na: if True (default) keep literal "NA" values in the sheet.
                 if False, use pandas default NA semantics (original behaviour).
    """
    non_mp_columns = [col for col in df.columns
                         if col.endswith(('EM', 'RA')) or col == 'estudiant']
    df = df.rename(columns=lambda col: col.split('_')[0] if col not in non_mp_columns else col)
    mp_grade_columns = [col for col in df.columns if not col.endswith('EM') and
                         not col.endswith('RA') and col != 'estudiant']
    df_without_mp_grades = df[non_mp_columns].copy()

    ra_codes = [col for col in df_without_mp_grades.columns if col != 'estudiant']

    mp_groups = {mp: [] for mp in mp_codes}
    sorted_mp_codes = sorted(mp_codes, key=len, reverse=True)
    for ra_code in ra_codes:
        for mp_code in sorted_mp_codes:
            if ra_code.startswith(mp_code + '_'):
                mp_groups[mp_code].append(ra_code)
                break

    new_columns = ['estudiant']
    for mp_code in mp_codes:
        for ra in mp_groups[mp_code]:
            new_columns.append(ra)
        if mp_code in mp_codes_with_em:
            new_columns.extend([
                f'{mp_code} CENTRE',
                f'{mp_code} EMPRESA',
                f'{mp_code}'
            ])
        else:
            new_columns.append(f'{mp_code}')

    export_df = df_without_mp_grades.copy()

    em_to_mp = {}
    for mp_code in mp_codes_with_em:
        em_codes = [col for col in export_df.columns if col.startswith(f'{mp_code}_') and col.endswith('EM')]
        for em_code in em_codes:
            em_to_mp[em_code] = mp_code

    for col_name in new_columns:
        if col_name not in export_df.columns:
            export_df[col_name] = pd.Series(dtype='float64', index=export_df.index)

    export_df = export_df.reindex(columns=new_columns)

    for mp_code in mp_grade_columns:
        if mp_code in export_df.columns:
            export_df[mp_code] = df[mp_code]

    # Add sequential index column
    export_df.insert(0, '#', range(1, len(export_df) + 1))

    output_path = output_path.replace('.csv', '.xlsx')
    export_df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    def get_col_letter(header):
        for cell in ws[1]:
            if str(cell.value).strip() == header:
                return get_column_letter(cell.column)
        return None

    if em_to_mp:
        for em_code, mp_code in em_to_mp.items():
            empresa_header = f'{mp_code} EMPRESA'
            em_col = get_col_letter(em_code)
            empresa_col = get_col_letter(empresa_header)
            if em_col and empresa_col:
                for row_idx in range(2, ws.max_row + 1):
                    em_cell = ws[f'{em_col}{row_idx}']
                    empresa_cell = ws[f'{empresa_col}{row_idx}']
                    if em_cell.value is not None:
                        empresa_cell.value = em_cell.value
                ws.delete_cols(ws[em_col][0].column, 1)
        wb.save(output_path)

    # Re-read the written file once. If preserve_na is True, instruct pandas
    # not to interpret default NA strings (so "NA" remains a literal string).
    export_df = pd.read_excel(output_path, keep_default_na=(not preserve_na))

    numeric_cols = export_df.select_dtypes(include=['int64', 'float64']).columns
    for col in numeric_cols:
        export_df[col] = pd.to_numeric(export_df[col], errors='coerce').astype('float64')

    non_numeric_cols = export_df.columns.difference(numeric_cols)
    export_df[non_numeric_cols] = export_df[non_numeric_cols].fillna('')

    export_df.to_excel(output_path, index=False)

    apply_row_formatting(output_path, mp_codes_with_em, mp_codes)
    apply_conditional_formatting(output_path, mp_groups, mp_codes_with_em, mp_codes, preserve_na=preserve_na)
    print(f"\t- Exported {len(export_df)} entries to {output_path}")
