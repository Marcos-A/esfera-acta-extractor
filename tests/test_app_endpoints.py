from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

import app as app_module


class ImmediateThread:
    def __init__(self, *, target, args=(), kwargs=None, daemon=None):
        self.target = target
        self.args = args
        self.kwargs = kwargs or {}
        self.daemon = daemon

    def start(self) -> None:
        self.target(*self.args, **self.kwargs)


def test_health_and_admin_login_smoke(client) -> None:
    health = client.get("/health")
    login = client.post(
        "/admin/login",
        data={"username": "admin", "password": "s3cret"},
        follow_redirects=False,
    )

    assert health.status_code == 200
    assert health.get_json() == {"status": "ok"}
    assert login.status_code == 302
    assert login.headers["Location"].endswith("/admin")


def test_convert_rejects_missing_or_invalid_uploads(client) -> None:
    missing = client.post("/convert", data={}, content_type="multipart/form-data")
    mixed = client.post(
        "/convert",
        data={
            "file": [
                (BytesIO(b"%PDF-1.4"), "first.pdf"),
                (BytesIO(b"zip"), "batch.zip"),
            ]
        },
        content_type="multipart/form-data",
    )

    assert missing.status_code == 400
    assert "Cal carregar almenys un fitxer" in missing.get_json()["error"]
    assert mixed.status_code == 400
    assert "tots han de ser PDF" in mixed.get_json()["error"]


def test_single_pdf_convert_flow_succeeds_with_stubbed_converter(app_instance, client, monkeypatch, tmp_path: Path) -> None:
    def fake_convert_pdf_to_excel(pdf_path: str | Path, output_dir: str | Path) -> Path:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = output_dir / "GROUP_A.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "#"
        worksheet["B1"] = "ESTUDIANT"
        worksheet["A2"] = 1
        worksheet["B2"] = "Alice Example"
        workbook.save(workbook_path)
        return workbook_path

    monkeypatch.setattr(app_module, "convert_pdf_to_excel", fake_convert_pdf_to_excel)
    monkeypatch.setattr(app_module.threading, "Thread", ImmediateThread)

    response = client.post(
        "/convert",
        data={"file": (BytesIO(b"%PDF-1.4 fake"), "sample.pdf")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 202
    payload = response.get_json()
    job_id = payload["job_id"]

    status_response = client.get(f"/convert/status/{job_id}")
    assert status_response.status_code == 200
    status_payload = status_response.get_json()
    assert status_payload["status"] == "success"
    assert status_payload["download_ready"] is True
    assert status_payload["returned_name"] == "GROUP_A.xlsx"

    job = app_instance.audit_store.get_job(job_id)
    metadata = app_module._load_metadata(job["metadata_json"])
    output_path = Path(metadata["output_path"])
    work_dir = Path(metadata["work_dir"])
    assert output_path.exists()
    assert work_dir.exists()

    download = client.get(f"/convert/download/{job_id}")
    assert download.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in download.headers["Content-Type"]

    reopened = load_workbook(BytesIO(download.data))
    assert reopened.active["B2"].value == "Alice Example"
    assert not output_path.exists()
    assert not work_dir.exists()
