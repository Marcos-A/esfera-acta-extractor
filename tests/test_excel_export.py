from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.excel_processor import export_excel_with_spacing


def test_export_excel_with_spacing_creates_expected_public_safe_workbook(tmp_path: Path, export_case) -> None:
    dataframe, mp_codes_with_em, mp_codes = export_case
    output_path = tmp_path / "grades.xlsx"

    export_excel_with_spacing(dataframe, str(output_path), mp_codes_with_em, mp_codes)

    workbook = load_workbook(output_path)
    worksheet = workbook.active

    assert worksheet.title == "Acta"
    assert worksheet.freeze_panes == "C2"
    assert worksheet["A1"].value == "#"
    assert worksheet["B1"].value == "ESTUDIANT"
    assert worksheet["C1"].value == "MP01_CF01_\n1RA"
    assert worksheet["D1"].value == "MP01_CF01_\n2RA"
    assert worksheet["E1"].value == "MP01 CENTRE"
    assert worksheet["F1"].value == "MP01 EMPRESA"
    assert worksheet["G1"].value == "MP01"
    assert worksheet["H1"].value == "MP02_CF02_\n1RA"
    assert worksheet["I1"].value == "MP02"

    assert worksheet["A2"].value == 1
    assert worksheet["B2"].value == "Alice Example"
    assert worksheet["F2"].value == 8.5
    assert worksheet["G2"].value == 8
    assert worksheet["H2"].value in ("", None)
    assert worksheet["I2"].value == 4

    assert worksheet["B5"].value == "MP amb estada a l'empresa"
    assert worksheet["B6"].value == "MP sense estada a l'empresa"
    assert len(worksheet.conditional_formatting) >= 7


def test_export_excel_output_can_be_reopened_after_write(tmp_path: Path, export_case) -> None:
    dataframe, mp_codes_with_em, mp_codes = export_case
    output_path = tmp_path / "roundtrip.xlsx"

    export_excel_with_spacing(dataframe, str(output_path), mp_codes_with_em, mp_codes)

    reopened = load_workbook(output_path, data_only=False)
    assert reopened.active.max_column == 9
    assert reopened.active.max_row >= 6



def test_export_excel_with_spacing_can_append_pending_ra_summary_sheet(tmp_path: Path, export_case) -> None:
    dataframe, mp_codes_with_em, mp_codes = export_case
    output_path = tmp_path / "summary-grades.xlsx"

    export_excel_with_spacing(
        dataframe,
        str(output_path),
        mp_codes_with_em,
        mp_codes,
        include_summary_sheet=True,
    )

    workbook = load_workbook(output_path)

    assert workbook.sheetnames == ["Acta", "Resum"]
    summary = workbook["Resum"]
    assert summary["A1"].value == "#"
    assert summary["B1"].value == "ESTUDIANT"
    assert summary["C1"].value == "RA PENDENTS"
    assert summary["D1"].value == "OBSERVACIONS"
    assert summary["B2"].value == "Alice Example"
    assert summary["C2"].value == "MP01:\n- RA2\n\nMP02:\n- RA1"
    assert summary["C3"].value in ("", None)
    assert summary.freeze_panes == "C2"
    assert summary.row_dimensions[2].height > summary.row_dimensions[3].height
    assert summary.row_dimensions[3].height == 24



def test_pending_ra_summary_uses_friendly_mp_name_when_available(tmp_path: Path) -> None:
    dataframe = __import__("pandas").DataFrame([
        {"estudiant": "Alice Example", "0647_CF01_1RA": "PDT", "0647_CF01_2RA": 7},
    ])

    export_excel_with_spacing(
        dataframe,
        str(tmp_path / "friendly-summary.xlsx"),
        mp_codes_with_em=[],
        mp_codes=["0647"],
        include_summary_sheet=True,
    )

    workbook = load_workbook(tmp_path / "friendly-summary.xlsx")
    summary = workbook["Resum"]
    assert summary["C2"].value == "0647. Gestió de la documentació jurídica i empresarial:\n- RA1"



def test_pending_ra_summary_sorts_double_digit_ras_numerically(tmp_path: Path) -> None:
    dataframe = __import__("pandas").DataFrame([
        {
            "estudiant": "Alice Example",
            "MP01_CF01_1RA": 7,
            "MP01_CF01_2RA": "PDT",
            "MP01_CF01_9RA": "PDT",
            "MP01_CF01_10RA": "PDT",
        },
    ])

    export_excel_with_spacing(
        dataframe,
        str(tmp_path / "sorted-summary.xlsx"),
        mp_codes_with_em=[],
        mp_codes=["MP01"],
        include_summary_sheet=True,
    )

    workbook = load_workbook(tmp_path / "sorted-summary.xlsx")
    summary = workbook["Resum"]
    assert summary["C2"].value == "MP01:\n- RA2\n- RA9\n- RA10"
